#!/usr/bin/env python

"""schema_generator.py: Main module to parse csv and generate and save SQL (Oracle) DDL Scripts"""

__author__ = "Cody Putnam (csp05)"
__version__ = "23.02.06.0"

import logging
import os
import csv
import time
import shutil
import copy
import openpyxl as xl
from openpyxl.comments import Comment
from concurrent.futures import ThreadPoolExecutor
from config import Config

threads = os.cpu_count() ^ 2
logger = logging
config = Config().config
# Update changes to CSV file in default_schema_row AND schema_headers
# Columns must be in file order
default_schema_row = { # Order of columns in file
                "schema": '',
                "table": '',
                "field": '',
                "type": '',
                "size": '',
                "units": '',
                "not_null": '',
                "primary_key": '',
                "default": '',
                "index": '',
                "sequence_start": '',
                "pop_by_trigger": '',
                "invisible": '',
                "virtual": '',
                "virtual_expr": '',
                "check_constraint": '',
                "lob_deduplication": '',
                "lob_compression": '',
                "lob_caching": '',
                "lob_logging": '',
                "fk_to_table": '',
                "fk_to_field": '',
                "gen_audit_columns": '',
                "gen_history_tables": '',
                "loader_package": '',
                "loader_parent_table": '',
                "table_comment": '',
                "column_comment": ''
                }
default_grant_row = { #Order of columns in file
                "schema": '',
                "table": '',
                "user": '',
                "insert": '',
                "update": '',
                "delete": ''
                }

default_grantexec_row = { #Order of columns in file
                "schema": '',
                "proc": '',
                "user": ''
                }

schema_headers = [
                'Schema',
                'Table',
                'Field',
                'Type',
                'Size',
                'Units',
                'Not Null',
                'Primary Key',
                'Default',
                'Index',
                'Sequence Start',
                'Pop by Trigger',
                'Invisible',
                'Virtual',
                'Virtual Expression',
                'Simple Check Constraint',
                'LOB Deduplication',
                'LOB Compression (LOW/MEDIUM/HIGH)',
                'LOB Caching',
                'LOB Logging',
                'FK to Table',
                'FK to Field',
                'Gen Audit Columns',
                'Gen History Table (Automated)',
                'Include in Loader Package',
                'Parent table for Loader',
                'Table Comment',
                'Column Comment'
                ]

grant_headers = [
                'Schema',
                'Table/View',
                'User',
                'Insert',
                'Update',
                'Delete'
                ]

grantexec_headers = [
                'Schema',
                'Package/Procedure/Function',
                'User'
                ]

schema_header_comments = {
                        'Schema':'Schema Name to be used',
                        'Table':'Table Name to be used',
                        'Field':'Field/Column Name to be used',
                        'Type':'Field type (e.g. VARCHAR2, NUMBER, DATE, etc.)',
                        'Size':'Field size, if applicable',
                        'Units':'Units for field size, if applicable (e.g. CHAR, BYTE)',
                        'Not Null':'Should field be marked as Not Null? Y if yes, else no',
                        'Primary Key':'Is this field the primary key of the table? Every table should have 1 primary key. If more than 1 is specified, a compound key is generated from all keys marked.',
                        'Default':'Default value for the field if value not provided on INSERT',
                        'Index':'Should field be indexed? Y = Indexed; U = Indexed as unique and constrained. Add a number to link indices to make compound indices (eg. all fields marked "U1" would make 1 compound index and constraint)',
                        'Sequence Start':'Add a sequence linked to the field. If a number is provided, this will be the first number in the sequence. If a sequence name is provided, it will reuse the sequence named instead.',
                        'Pop by Trigger':'Should field be auto-populated from sequence by trigger? Y if yes, else no',
                        'Invisible':'Should field be hidden from select * queries? Y if yes, else no',
                        'Virtual':'Should field be a virtual field? Y if yes, else no. Define the expression to be used in the next field.',
                        'Virtual Expression':'Expression to evaluate for the fields value, if Virtual = Y.',
                        'Simple Check Constraint':'Formula for a basic Check Constraint for restricting allowed values in a field.',
                        'LOB Deduplication':'If field is a CLOB or BLOB type field, should deduplication be applied.',
                        'LOB Compression (LOW/MEDIUM/HIGH)':'If field is a CLOB or BLOB type field, what level of storage compression should be used.',
                        'LOB Caching':'If field is a CLOB or BLOB type field, should caching be used.',
                        'LOB Logging':'If field is a CLOB or BLOB type field, should logging be enabled.',
                        'FK to Table':'Foreign Key: Define the table to link field back to.',
                        'FK to Field':'Foreign Key: Define the field to link field back to.',
                        'Gen Audit Columns':'Should audit columns (u_name, u_date) be injected into this table. Also creates triggers to auto-populate these fields. Only applicable on first line of given table.',
                        'Gen History Table (Automated)':'Should a history table be generated along with appropriate triggers to populate it. Only applicable on first line of given table.',
                        'Include in Loader Package': 'Should table be included in a loader package for INSERT, UPDATE, DELETE statements. Value should either be Y for a generically named package or should be the name of the package. Only applicable on the first line of given table.',
                        'Parent table for Loader': 'If applicable, should be the name of the parent table to the given table. MUST BE FROM SAME SCHEMA DEFINITION FILE AND ALSO LOADED. Only applicable on the first line of given table.',
                        'Table Comment':'Table level comment. Only applicable on first line of given table.',
                        'Column Comment':'Field comment',
                        }
grant_header_comments = {
                        'Schema':'Schema Name to be used',
                        'Table/View':'Table or View Name to be used',
                        'User':'Username or Role to be assigned permission',
                        'Insert':'Add X in column if user/role should be assigned this permission. SELECT permission is assumed if ID is included at all',
                        'Update':'Add X in column if user/role should be assigned this permission. SELECT permission is assumed if ID is included at all',
                        'Delete':'Add X in column if user/role should be assigned this permission. SELECT permission is assumed if ID is included at all'
                        }
grantexec_header_comments = {
                        'Schema':'Schema Name to be used',
                        'Package/Procedure/Function':'Executable Name to be used',
                        'User':'Username or Role to be assigned EXECUTE permission'
                        }

import sql_script as sql
import build_script as build
import clean_script as clean
from table import Table, Column

# Logging settings for overall project. Loads logging level to use from config file (config.py -> conf.json)
logger.basicConfig(level= logging.getLevelName(config['logging']['level']['setting']), 
                    filename='schema_generator.log', 
                    format='%(asctime)s - %(levelname)s - %(message)s', 
                    datefmt='%Y-%m-%d %H:%M:%S'
                    )

# Load settings from config file (config.py -> conf.json)
outputDir = config['files']['output-directory']['setting']
fileType = config['files']['schema-file-type']['setting']
csvfile = config['files']['schema-file-csv']['setting']
grantsFile = config['files']['grants-file-csv']['setting']
xlsxFile = config['files']['schema-file']['setting']

sortColumns = config['sorting']['columns-nullable']['setting']

history_package = config['history-tables']['use-procedures']['setting']

loader_package = config['loader-package']['enable']['setting']

doClean = config['clean-script']['setting']

# Dictionary to hold tables to process
tables = {}
tableCount: int

todo = []
todoGrants = []
todoGrantsExecute = []

addGrants = []
addGrantsExec = []

def mergeListtoString(toMerge: list) -> str:
    """
    mergeListtoString(toMerge)

    Takes a provided list and converts it into a single string that is comma-delimited
 
    Parameters:
        csvrow: tuple
            Unprocessed row from formatted CSV file, addressed by numeric index

    Returns:
        str
            String of all the contents of the provided list in a comma-delimited format
    """

    outString = ''
    for i in toMerge:
        outString = f'{outString}{i},'
    return outString.rstrip(',')


def convertToDict(csvrow: tuple, grantFile: bool) -> dict:
    """
    convertToDict(csvrow, grantFile)

    Takes line from formatted CSV file and converts it to a key-linked dictionary
    Allows for adding (/ removing) columns from CSV document in future and having one place to define new field to use
 
    Parameters:
        csvrow: tuple
            Unprocessed row from formatted CSV file, addressed by numeric index
        grantFile: bool
            If the file is of type Grants, set to True to process appropriately.
            Otherwise will be processed as Schema type file

    Returns:
        dict
            Dictionary containing fields parsed into keys from the supplied row out of the CSV
    """

    d = {}
    if grantFile:
        d['schema'] = csvrow[0].strip()
        d['table'] = csvrow[1].strip()
        d['user'] = csvrow[2].strip()
        d['insert'] = csvrow[3].strip()
        d['update'] = csvrow[4].strip()
        d['delete'] = csvrow[5].strip()
    
    # Process as Schema type file
    else:
        # Populate default_schema_row template in file order
        d = copy.deepcopy(default_schema_row)
        index = 0
        try:
            for key in d.keys():
                d[key] = csvrow[index].strip()
                index += 1
        except:
            logger.error(f'CSV File is missing columns. Columns expected: {schema_headers}')
            raise IndexError
    
    # Return dict row
    return d


def csvRead(filename: str, grantFile: bool = False):
    """
    csvRead(filename, grantFile)

    Reads a formatted CSV file and parses out the data. 
    Skips blank lines or header rows (so these can and *should* be left in the file)
 
    Parameters:
        filename: str
            Path string of file to be processed
        grantFile: bool, default = False
            If the file is of type Grants, set to True to process appropriately.
            Otherwise will be processed as Schema type file

    Returns:
        list
            List of dictionaries (one per row) in CSV file
    """

    todotemp = []

    # If file defined is not found, generate a new copy with the appropriate headers
    if not os.path.exists(filename):
        logger.info('Generating file...')
        with open(filename, 'w') as file:
            if grantFile:
                file.write(mergeListtoString(grant_headers))
            else:
                file.write(mergeListtoString(schema_headers))
    
    # If file is found, process rows
    else:
        with open(filename, newline='') as file:
            reader = csv.reader(file)
            logger.info('Appending rows to ToDo list')
            firstrow = True
            for row in reader:
                
                #skip header row or blank rows (determined if first cell (Schema) is blank)
                if not firstrow and row[0] not in ['Schema', '']:
    
                    # Convert row to dictionary to allow key-based lookup of values
                    rowdict = convertToDict(row, grantFile)
                    if grantFile:
                        todoGrants.append(rowdict)
                    else:
                        todo.append(rowdict)
                elif firstrow:
                    firstrow = False
    # If file not found and generated, 'todo' will be empty list and have nothing to process


def xlsxRead(filename: str):
    """
    xlsxRead(filename)

    Reads an XLSX file and parses out the data. 
    Skips blank lines or header rows (so these can and *should* be left in the file)
 
    Parameters:
        filename: str
            Path string of file to be processed
    """
    
    # If file defined is not found, generate a new copy with the appropriate headers
    if not os.path.exists(filename):
        logger.info('Generating xlsx file...')
        wb = xl.Workbook()
        schema_sheet = wb.active
        schema_sheet.title = 'Schema Design'
        schema_sheet.sheet_properties.tabColor = '009900'
        grants_sheet = wb.create_sheet('Grants')
        grants_sheet.sheet_properties.tabColor = '000099'
        grantsexec_sheet = wb.create_sheet('Grants - Execute')
        grantsexec_sheet.sheet_properties.tabColor = '009999'

        for i in range(0, len(grant_headers)):
            cell = grants_sheet.cell(row=1, column=i+1)
            if cell.value == None: 
                cell.value = grant_headers[i]
                cell.style = 'Headline 3'
                try:
                    cell.comment = Comment(grant_header_comments[grant_headers[i]], __author__, 75, 500)
                except:
                    logger.warning(f'No comment defined for Grant header: {grant_headers[i]}')

        for i in range(0, len(grantexec_headers)):
            cell = grantsexec_sheet.cell(row=1, column=i+1)
            if cell.value == None: 
                cell.value = grantexec_headers[i]
                cell.style = 'Headline 3'
                try:
                    cell.comment = Comment(grantexec_header_comments[grantexec_headers[i]], __author__, 75, 500)
                except:
                    logger.warning(f'No comment defined for Grants - Execute header: {grantexec_headers[i]}')

        for i in range(0, len(schema_headers)):
            cell = schema_sheet.cell(row=1, column=i+1)
            if cell.value == None: 
                cell.value = schema_headers[i]
                cell.style = 'Headline 3'
                try:
                    cell.comment = Comment(schema_header_comments[schema_headers[i]], __author__, 75, 500)
                except:
                    logger.warning(f'No comment defined for Schema header: {schema_headers[i]}')
        
        grants_sheet.freeze_panes = grants_sheet['A2']
        grantsexec_sheet.freeze_panes = grantsexec_sheet['A2']
        schema_sheet.freeze_panes = schema_sheet['D2']

        try:
            wb.save(filename)
        except:
            logger.error('Unable to save XLSX workbook. May be already open in another process.')
        wb.close()

    # If file is found, process rows
    else:
        can_load = False
        # Load xlsx Workbook
        try:
            wb = xl.load_workbook(filename)
            can_load = True
        except:
            logger.error('Unable to load XLSX workbook. May be already open in another process.')
        if can_load:
            try:
                schema_sheet = wb['Schema Design']
            except:
                schema_sheet= wb.create_sheet('Schema Design')
                schema_sheet.sheet_properties.tabColor = '009900'
                schema_sheet.freeze_panes = schema_sheet['D2']
            try:
                grants_sheet = wb['Grants']
            except:
                grants_sheet = wb.create_sheet('Grants')
                grants_sheet.sheet_properties.tabColor = '000099'
                grants_sheet.freeze_panes = grants_sheet['A2']
            try:
                grantsexec_sheet = wb['Grants - Execute']
            except:
                grantsexec_sheet = wb.create_sheet('Grants - Execute')
                grantsexec_sheet.sheet_properties.tabColor = '009999'
                grantsexec_sheet.freeze_panes = grantsexec_sheet['A2']
            is_updated = False
            # Check Schema column headers, insert any new columns needed
            tot_schema_header = len(schema_headers)
            for i in range(0, tot_schema_header):
                cell = schema_sheet.cell(row=1, column=i+1)
                if cell.value != schema_headers[i]:
                    logger.debug(f'Missing Schema header: {schema_headers[i]}')
                    schema_sheet.insert_cols(i+1)
                    cell = schema_sheet.cell(row=1, column=i+1)
                    cell.value = schema_headers[i]
                    cell.style = 'Headline 3'
                    try:
                        cell.comment = Comment(schema_header_comments[schema_headers[i]], __author__, 75, 500)
                    except:
                        logger.warning(f'No comment defined for Schema header: {schema_headers[i]}')
                    is_updated = True
            # Are there extra columns at the end of the sheet? If so, delete them
            extra_cols = schema_sheet.max_column > tot_schema_header
            logger.debug(f'Extra columns found in Schema sheet: {extra_cols}')
            if extra_cols:
                to_delete = schema_sheet.max_column - tot_schema_header
                logger.warning(f'Deleting {to_delete} columns from Schema Design worksheet')
                schema_sheet.delete_cols(tot_schema_header + 1, to_delete)
                is_updated = True
            
            # Check Grant column headers, insert any new columns needed
            tot_grant_header = len(grant_headers)
            for i in range(0, tot_grant_header):
                cell = grants_sheet.cell(row=1, column=i+1)
                if cell.value != grant_headers[i]:
                    logger.debug(f'Missing Grant header: {grant_headers[i]}')
                    grants_sheet.insert_cols(i+1)
                    cell = grants_sheet.cell(row=1, column=i+1)
                    cell.value = grant_headers[i]
                    cell.style = 'Headline 3'
                    try:
                        cell.comment = Comment(grant_header_comments[grant_headers[i]], __author__, 75, 500)
                    except:
                        logger.warning(f'No comment defined for Grant header: {grant_headers[i]}')
                    is_updated = True
            # Are there extra columns at the end of the sheet? If so, delete them
            extra_cols = grants_sheet.max_column > tot_grant_header
            logger.debug(f'Extra columns found in Grants sheet: {extra_cols}')
            if extra_cols:
                to_delete = grants_sheet.max_column - tot_grant_header
                logger.warning(f'Deleting {to_delete} columns from Grants worksheet')
                grants_sheet.delete_cols(tot_grant_header + 1, to_delete)
                is_updated = True
            
            # Check Grant column headers, insert any new columns needed
            tot_grantexec_header = len(grantexec_headers)
            for i in range(0, tot_grantexec_header):
                cell = grantsexec_sheet.cell(row=1, column=i+1)
                if cell.value != grantexec_headers[i]:
                    logger.debug(f'Missing Grants: Execute header: {grantexec_headers[i]}')
                    grantsexec_sheet.insert_cols(i+1)
                    cell = grantsexec_sheet.cell(row=1, column=i+1)
                    cell.value = grantexec_headers[i]
                    cell.style = 'Headline 3'
                    try:
                        cell.comment = Comment(grantexec_header_comments[grantexec_headers[i]], __author__, 75, 500)
                    except:
                        logger.warning(f'No comment defined for Grants: Execute header: {grantexec_headers[i]}')
                    is_updated = True
            # Are there extra columns at the end of the sheet? If so, delete them
            extra_cols = grantsexec_sheet.max_column > tot_grantexec_header
            logger.debug(f'Extra columns found in Grants: Execute sheet: {extra_cols}')
            if extra_cols:
                to_delete = grantsexec_sheet.max_column - tot_grantexec_header
                logger.warning(f'Deleting {to_delete} columns from Grants worksheet')
                grantsexec_sheet.delete_cols(tot_grantexec_header + 1, to_delete)
                is_updated = True
            
            if is_updated:
                logger.info('XLSX Workbook modified. Attempting to save changes...')
                try:
                    wb.save(filename)
                    logger.info('XLSX Workbook saved successfully!')
                except:
                    logger.error('Unable to save XLSX workbook. May be already open in another process.')

            # Read contents into dictionary (Schema)
            holdrows = schema_sheet.iter_rows(min_row=2, values_only=True)
            for row in holdrows:
                if row[0] != None:
                    row_work = copy.deepcopy(default_schema_row)
                    for (index, key) in enumerate(row_work):
                        xlsxvalue = row[index]
                        if xlsxvalue == None:
                            row_work[key] = ''
                        else:
                            row_work[key] = str(xlsxvalue)
                    todo.append(row_work)

            # Read contents into dictionary (Grants)
            holdrows = grants_sheet.iter_rows(min_row=2, values_only=True)
            for row in holdrows:
                if row[0] != None:
                    row_work = copy.deepcopy(default_grant_row)
                    for (index, key) in enumerate(row_work):
                        xlsxvalue = row[index]
                        if xlsxvalue == None:
                            row_work[key] = ''
                        else:
                            row_work[key] = str(xlsxvalue)
                    todoGrants.append(row_work)

            # Read contents into dictionary (Grants)
            holdrows = grantsexec_sheet.iter_rows(min_row=2, values_only=True)
            for row in holdrows:
                if row[0] != None:
                    row_work = copy.deepcopy(default_grantexec_row)
                    for (index, key) in enumerate(row_work):
                        xlsxvalue = row[index]
                        if xlsxvalue == None:
                            row_work[key] = ''
                        else:
                            row_work[key] = str(xlsxvalue)
                    todoGrantsExecute.append(row_work)

            wb.close()


def writeGrants(filename: str):
    """
    writeGrants(filename)

    Writes to an XLSX file, adding basic grants for all known tables
 
    Parameters:
        filename: str
            Path string of file to be processed
    """

    # If file is found, process rows
    can_load = False
    # Load xlsx Workbook
    try:
        wb = xl.load_workbook(filename)
        can_load = True
    except:
        logger.error('Unable to load XLSX workbook. May be already open in another process.')
    if can_load:
        try:
            grants_sheet = wb['Grants']
        except:
            grants_sheet = wb.create_sheet('Grants')
            grants_sheet.sheet_properties.tabColor = '000099'
            grants_sheet.freeze_panes = grants_sheet['A2']
        row_num = 2
        for grant in addGrants:
            i = 1
            for key in grant.keys():
                cell = grants_sheet.cell(row=row_num, column=i)
                cell.value = grant[key]
                i += 1
            todoGrants.append(grant)
            row_num += 2
        try:
            wb.save(filename)
            logger.info('XLSX Workbook saved successfully!')
        except:
            logger.error('Unable to save XLSX workbook. May be already open in another process.')
        wb.close()


def writeGrantsExec(filename: str):
    """
    writeGrants(filename)

    Writes to an XLSX file, adding basic grants for all known tables
 
    Parameters:
        filename: str
            Path string of file to be processed
    """

    # If file is found, process rows
    can_load = False
    # Load xlsx Workbook
    try:
        wb = xl.load_workbook(filename)
        can_load = True
    except:
        logger.error('Unable to load XLSX workbook. May be already open in another process.')
    if can_load:
        try:
            grantsexec_sheet = wb['Grants - Execute']
        except:
            grantsexec_sheet = wb.create_sheet('Grants - Execute')
            grantsexec_sheet.sheet_properties.tabColor = '009999'
            grantsexec_sheet.freeze_panes = grantsexec_sheet['A2']
        row_num = 2
        for grant in addGrantsExec:
            i = 1
            for key in grant.keys():
                cell = grantsexec_sheet.cell(row=row_num, column=i)
                cell.value = grant[key]
                i += 1
            todoGrantsExecute.append(grant)
            row_num += 2
        try:
            wb.save(filename)
            logger.info('XLSX Workbook saved successfully!')
        except:
            logger.error('Unable to save XLSX workbook. May be already open in another process.')
        wb.close()


def createGrant(schema: str, tablename: str, level: tuple):
    app_account = schema.upper().split('_OWNER')[0]
    addGrants.append({"schema": schema, 
                      "table": tablename, 
                      "user": app_account,
                      "insert": level[0],
                      "update": level[1],
                      "delete": level[2]
                      })


def createGrantHistory(schema: str):
    app_account = schema.upper().split('_OWNER')[0]
    addGrantsExec.append({"schema": schema, 
                          "proc": f'{app_account}_HISTORY', 
                          "user": app_account
                          })


def createGrantLoader(schema: str):
    app_account = schema.upper().split('_OWNER')[0]
    addGrantsExec.append((schema, f'{app_account}_LOADER', app_account))


def generateHistoryTables(table: Table, tableNum: int):
    schematable = f'{table.schema}.H_{table.name}'
    histTable = table.genHistoryTable(tableNum)
    if histTable is not None:
        tables[schematable] = histTable


def processTable(table: Table):
    """
    processTable(table)

    Processes a given table and generates all the scripts for that table. 
    Also iterates over all columns and generates any column-level scripts

    Parameters:
        table: Table
            Table object to be processed
    """

    logger.info(f'Processing table {table.schema}.{table.name}...')
    
    # Process compound indexes and remove any invalid entries
    table.cleanCompoundIndex()

    # If table needs to be included in Loader package, do this now
    if table.needsloader:
        tables = [table]
        localtable = table
        while localtable.loaderParent != None:
            localtable = localtable.loaderParent
            tables.append(localtable)
        sql.addToLoaderPackage(table.schema, tables)

    # If Gen History Table is True, create history table object and process for scripts
    if table.needshistory:
        logger.info('History Table and Structure requested.')
        sql.writeHistoryTriggers(table.schema, table.name, table.genColumnListForHistory())
    
    # If Gen Audit Columns is True, inject audit columns into table
    if table.needsaudit:
        logger.info('Injecting audit columns...')
        table.genAuditColumns()
        sql.writeAuditTrigger(table.schema, table.name)

    # If sorting is enabled via config, do so now
    if sortColumns:
        table.sortColumns()

    # If table has a comment, add to comment queue (written at end)
    if table.comment != '':
        sql.addTableComment(table.schema, table.name, table.comment)
    
    # If table has multiple PKs defined, create compound key scripts
    if table.hasCompoundPK():
        logger.debug(f'Table {table.name} has a compound primary key')
        sql.writeCompoundIndexScript(table.schema, 
                                    table.name, 
                                    table.tablespace, 
                                    table.getPKFields(), 
                                    table.tableNumber)
    
    # If table has any compound indexes defined (after the cleanup), create scripts
    if table.hasCompoundIndex():
        logger.debug(f'Table {table.name} has one or more compound indexes')
        indexFields = table.getIndexFields()
        for key in indexFields.keys():
            sql.writeCompoundIndexScript(table.schema, 
                                        table.name, 
                                        table.tablespace, 
                                        indexFields[key], 
                                        table.tableNumber, 
                                        table.indexcount, 
                                        primaryKey = False, 
                                        unique = table.isCompoundIndexUnique(key))
            table.indexcount += 1

    # Write table script to file
    sql.writeTableScript(table.schema, table.name, table.genColumnList(), table.tablespace)

    if table.ishistory:
        level = ('','','')
    else:
        level = ('X','X','X')
    createGrant(table.schema, table.name, level)

    # Process all columns individually (cannot be multi-threaded due to shared index and FK counts per table)
    for col in table.columns:
        processColumn(table, col)


def processColumn(table: Table, column: Column):
    """
    processColumn(table, column)

    Processes a given column and generates all the scripts for that column. 

    Parameters:
        table: Table
            Table object to be processed
        column: Column
            Column object to be processed
    """

    logger.debug(f'Writing scripts for {table.name}.{column.field}')
    # If table does not have compound PK and column is marked as PK, generate scripts for PK
    if column.primarykey and not table.hasCompoundPK():
        logger.debug(f'{table.name}.{column.field} is Primary Key')
        sql.writeIndexScript(table.schema, 
                            table.name, 
                            table.tablespace, 
                            column.field, 
                            table.tableNumber, 
                            primaryKey = column.primarykey)
    # If not PK, but marked as indexed, check if in compound index then process if not
    elif column.indexed: ### and not table.isFieldInCompoundIndex(column.field): ### Columns no longer marked as "indexed" unless it's an individual index
        logger.debug(f'{table.name}.{column.field} is Indexed and unique = {column.unique}')
        sql.writeIndexScript(table.schema, 
                            table.name, 
                            table.tablespace, 
                            column.field, 
                            table.tableNumber, 
                            table.indexcount, 
                            primaryKey = False,
                            unique = column.unique)
        table.indexcount += 1
    
    # If column needs a sequence, generate scripts 
        # If column needs a sequence, generate scripts 
    # If column needs a sequence, generate scripts 
    # Flag for if sequence needs to be populated by trigger and generate that script as well
    if column.sequenced:
        if column.sequencetouse == None:
            logger.debug(f'{table.name}.{column.field} has an assigned Sequence. Trigger-fired = {column.triggered}')
            sql.writeSequenceScript(table.schema, 
                                    table.name, 
                                    column.field, 
                                    column.sequencestart, 
                                    column.triggered)
        # If reusing sequence, nothing to do unless trigger population is requested
        elif column.triggered:
            logger.debug(f'{table.name}.{column.field} is re-using sequence {column.sequencetouse}. Trigger-fired = {column.triggered}')
            sql.writeTriggerScript(table.schema, table.name, column.field, column.sequencetouse)
    
    # If column has a foreign key source table defined, generate FK scripts
    # Only need to check for FK Table because field is implied due to column loading logic
    if column.fksourcetable != None:
        logger.debug(f'{table.name}.{column.field} has a foreign key relation to {column.fksourcetable}.{column.fksourcefield}')
        sql.writeFKConstraintScript(table.schema, 
                                    column.fksourcetable, 
                                    column.fksourcefield, 
                                    table.name, 
                                    column.field, 
                                    table.tableNumber, 
                                    table.fkcount)
        table.fkcount += 1
    
    # If column has a simple Check Constraint defined, generate the script for this
    if column.checkconstraint != None:
        logger.debug(f'{table.name}.{column.field} has a simple Check Constraint defined as "{column.field} {column.checkconstraint}"')
        sql.writeSimpleCheckConstraintScript(table.schema,
                                            table.name,
                                            column.field,
                                            column.checkconstraint,
                                            table.tableNumber)
    
    # If column has a comment, add to comment queue to be written at end
    if column.comment != None:
        logger.debug(f'{table.name}.{column.field} has a comment')
        sql.addColumnComment(table.schema, table.name, column.field, column.comment)


def addGrant(grant):
    level = 'S'
    if grant["insert"].upper() == 'X':
        level = f'{level}I'
    if grant["update"].upper() == 'X':
        level = f'{level}U'
    if grant["delete"].upper() == 'X':
        level = f'{level}D'
    logger.debug(f'Adding GRANT of {level} to {grant["user"]} on {grant["schema"]}.{grant["table"]}')
    sql.addGrant(grant["schema"], grant["table"], grant["user"], level)

def addGrantExec(grant):
    logger.debug(f'Adding GRANT of EXECUTE to {grant["user"]} on {grant["schema"]}.{grant["proc"]}')
    sql.addGrantExec(grant["schema"], grant["proc"], grant["user"])


def main():
    """
    main()
    
    Main application runtime
    """

    cleandir = False
    if os.getcwd() == os.path.realpath(outputDir):
        logger.error(f'Unable to write to working directory. Please specify a subfolder such as ".\\output"')
    elif os.path.exists(outputDir):
        # Try to delete output directory
        try: 
            shutil.rmtree(outputDir)
            os.makedirs(outputDir)
            cleandir = True
        # Log error and do not update cleandir
        except OSError as e: 
            logger.error(f'Error deleting output directory: {e.strerror}')
    # No output directory, create and continue
    else:
        os.makedirs(outputDir)
        cleandir = True
    # If successfully deleted and recreated the output directory, continue
    if cleandir: 
        # Read Schema CSV (grantFile = False default invoked)
        if fileType.lower() == 'csv':
            csvRead(csvfile)
        elif fileType.lower() == 'xlsx':
            xlsxRead(xlsxFile)
        else:
            logger.error(f'Invalid file type defined in config: {fileType}')
        tableCount = 1
        for row in todo:
            schematable = f'{row["schema"]}.{row["table"]}'
            logger.debug(f'Loading {schematable}.{row["field"]}')
            # If schema.table already in table dict: add column
            if schematable in tables.keys():
                newcol = Column()
                newcol.load(row)
                tables[schematable].addColumn(newcol)
            # If schema.table not in table dict: add new Table, then add column
            else:
                logger.debug(f'New table: {schematable}')
                if row["loader_parent_table"] != '':
                    loaderParent = tables[f'{row["schema"]}.{row["loader_parent_table"]}']
                else:
                    loaderParent = None
                tables[schematable] = Table(row["schema"], 
                                            row["table"], 
                                            row["gen_audit_columns"], 
                                            row["gen_history_tables"], 
                                            row["loader_package"],
                                            row["table_comment"],
                                            tableCount,
                                            loaderParent=loaderParent)
                tableCount += 1
                newcol = Column()
                newcol.load(row)
                tables[schematable].addColumn(newcol)
        logger.info('All tables and fields loaded')
        
        holdTables = copy.deepcopy(tables)
        # Extract and create History Tables
        for key in holdTables.keys():
            generateHistoryTables(holdTables[key], tableCount)
            tableCount += 1
        # Process each table and generate scripts
        with ThreadPoolExecutor(threads) as pool:
            pool.map(processTable, (tables[key] for key in tables.keys()))
        #for key in tables.keys():
        #    processTable(tables[key])

        # If 'use-procedure' setting is on, complete history package and write to file
        if history_package:
            logger.info('Saving history writing procedures to package')
            schemas = sql.writeHistoryPackage()
            for schema in schemas:
                createGrantHistory(schema)

        if loader_package:
            logger.info('Saving Loader package to file')
            schemas = sql.writeLoaderPackage()
            for schema in schemas:
                createGrantLoader(schema)

        # Write comments to file
        logger.info('Saving Comments for all tables and columns')
        sql.writeComments()

        if todoGrants == []:
            writeGrants(xlsxFile)
        
        if todoGrantsExecute == []:
            writeGrantsExec(xlsxFile)

        # Read Grants CSV and process contents
        if fileType.lower() == 'csv':
            csvRead(grantsFile, True)
        with ThreadPoolExecutor(threads) as pool:
            pool.map(addGrant, todoGrants)
        logger.info('Saving Grants for all tables')
        sql.writeGrants()

        # Write EXECUTE Grants
        with ThreadPoolExecutor(threads) as pool:
            pool.map(addGrantExec, todoGrantsExecute)
        logger.info('Saving Grants for all Executables')
        sql.writeGrantsExec()

        # Generate build.sql script in root of output
        logger.info('Writing final scripts to directory')
        build.genBuildScript()
        # If 'clean-script' setting is enabled, generate clean.sql script in root of output
        if doClean:
            clean.genCleanScript()


if __name__ == '__main__':
    logger.info('Starting application')
    starttime = time.perf_counter()
    main()
    endtime = time.perf_counter()
    close_message = f'All operations completed in {endtime - starttime:0.4f} seconds'
    logger.info(close_message)
    print(close_message)